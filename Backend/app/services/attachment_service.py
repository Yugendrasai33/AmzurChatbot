import base64
import csv
import io
import mimetypes
import uuid
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.attachment import Attachment

# Lazy imports for extraction libraries to avoid import errors if not installed
_CATEGORY_MAP: dict[str, str] = {
    "image/jpeg": "image",
    "image/png": "image",
    "image/gif": "image",
    "image/webp": "image",
    "video/mp4": "video",
    "video/webm": "video",
    "video/quicktime": "video",
    "application/pdf": "pdf",
    "text/plain": "code",
    "text/x-python": "code",
    "text/javascript": "code",
    "text/html": "code",
    "text/css": "code",
    "text/markdown": "code",
    "application/json": "code",
    "text/csv": "table",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "table",
    "application/vnd.ms-excel": "table",
    "application/msword": "document",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "document",
}


def get_allowed_mimes() -> list[str]:
    """Parse allowed MIME types from settings."""
    return [m.strip() for m in settings.ALLOWED_UPLOAD_MIMES.split(",") if m.strip()]


def classify_mime(mime_type: str) -> str | None:
    """Return the type_category for a given MIME type, or None if not allowed."""
    allowed = get_allowed_mimes()
    if mime_type not in allowed:
        return None
    return _CATEGORY_MAP.get(mime_type)


def _normalize_mime(mime_type: str | None) -> str | None:
    if not mime_type:
        return None
    normalized = mime_type.strip().lower()
    if normalized == "image/jpg":
        return "image/jpeg"
    return normalized


def detect_mime(file_bytes: bytes, filename: str | None = None, content_type: str | None = None) -> str:
    """Detect MIME type, preferring python-magic and falling back safely on Windows."""
    # Best effort: use libmagic when available.
    try:
        import magic

        detected = _normalize_mime(magic.from_buffer(file_bytes[:2048], mime=True))
        if detected:
            return detected
    except Exception:
        # libmagic is often unavailable on Windows dev machines.
        pass

    # Fall back to browser-provided content type if present.
    hinted = _normalize_mime(content_type)
    if hinted and hinted != "application/octet-stream":
        return hinted

    # Last fallback: infer from file extension.
    guessed = _normalize_mime(mimetypes.guess_type(filename or "")[0])
    if guessed:
        return guessed

    return "application/octet-stream"


def _sanitize_filename(filename: str) -> str:
    """Strip path traversal characters and return safe basename."""
    name = Path(filename).name
    # Remove any remaining path separators or null bytes
    name = name.replace("\x00", "").replace("/", "").replace("\\", "")
    return name or "unnamed"


def _get_extension(mime_type: str) -> str:
    """Get a safe file extension from MIME type."""
    ext = mimetypes.guess_extension(mime_type) or ""
    # mimetypes can return odd things like .ksh for text/plain
    overrides = {
        "text/plain": ".txt",
        "text/x-python": ".py",
        "text/javascript": ".js",
        "text/css": ".css",
        "text/html": ".html",
        "text/markdown": ".md",
        "text/csv": ".csv",
        "application/json": ".json",
        "application/msword": ".doc",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.ms-excel": ".xls",
        "video/quicktime": ".mov",
    }
    return overrides.get(mime_type, ext)


async def save_upload(
    db: AsyncSession,
    user_id: str,
    filename: str,
    file_bytes: bytes,
    detected_mime: str,
    type_category: str,
) -> Attachment:
    """Save an uploaded file to disk and create the DB record."""
    safe_name = _sanitize_filename(filename)
    ext = _get_extension(detected_mime)
    stored_name = f"{uuid.uuid4()}{ext}"

    # Create user directory
    user_dir = Path(settings.UPLOAD_DIR) / user_id
    user_dir.mkdir(parents=True, exist_ok=True)

    # Write file
    file_path = user_dir / stored_name
    file_path.write_bytes(file_bytes)

    # Relative path for DB storage
    relative_path = f"{user_id}/{stored_name}"

    attachment = Attachment(
        user_id=uuid.UUID(user_id),
        filename=safe_name,
        stored_path=relative_path,
        mime_type=detected_mime,
        size_bytes=len(file_bytes),
        type_category=type_category,
    )
    db.add(attachment)
    await db.flush()
    await db.refresh(attachment)
    return attachment


async def get_attachment(db: AsyncSession, file_id: str, user_id: str) -> Attachment | None:
    """Fetch an attachment only if owned by the user."""
    query = select(Attachment).where(
        Attachment.id == uuid.UUID(file_id),
        Attachment.user_id == uuid.UUID(user_id),
    )
    result = await db.execute(query)
    return result.scalar_one_or_none()


async def get_attachments_by_ids(
    db: AsyncSession, attachment_ids: list[str], user_id: str
) -> list[Attachment]:
    """Fetch multiple attachments owned by user."""
    if not attachment_ids:
        return []
    uuids = [uuid.UUID(aid) for aid in attachment_ids]
    query = select(Attachment).where(
        Attachment.id.in_(uuids),
        Attachment.user_id == uuid.UUID(user_id),
    )
    result = await db.execute(query)
    return list(result.scalars().all())


async def link_attachments_to_message(
    db: AsyncSession, attachment_ids: list[str], message_id: uuid.UUID
) -> None:
    """Set message_id on attachments after message is persisted."""
    if not attachment_ids:
        return
    uuids = [uuid.UUID(aid) for aid in attachment_ids]
    query = select(Attachment).where(Attachment.id.in_(uuids))
    result = await db.execute(query)
    for att in result.scalars().all():
        att.message_id = message_id


def get_file_path(attachment: Attachment) -> Path:
    """Get the absolute file path for an attachment."""
    return Path(settings.UPLOAD_DIR) / attachment.stored_path


def attachment_to_base64_url(attachment: Attachment) -> str:
    """Read an image attachment and return a base64 data URL."""
    file_path = get_file_path(attachment)
    data = file_path.read_bytes()
    b64 = base64.b64encode(data).decode("ascii")
    return f"data:{attachment.mime_type};base64,{b64}"


def extract_text_content(attachment: Attachment) -> str:
    """Extract text content from code, PDF, or table attachments."""
    file_path = get_file_path(attachment)

    if attachment.type_category == "code":
        return _extract_code(file_path)
    elif attachment.type_category == "pdf":
        return _extract_pdf(file_path)
    elif attachment.type_category == "table":
        return _extract_table(file_path, attachment.mime_type)
    elif attachment.type_category == "document":
        return _extract_document(file_path, attachment.mime_type)
    return ""


def _extract_code(file_path: Path) -> str:
    """Read raw code file content (limit to 10KB)."""
    text = file_path.read_text(encoding="utf-8", errors="replace")
    return text[:10000]


def _extract_pdf(file_path: Path) -> str:
    """Extract text from PDF using PyPDF2."""
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(str(file_path))
        pages_text = []
        for page in reader.pages[:20]:  # Limit to first 20 pages
            text = page.extract_text()
            if text:
                pages_text.append(text)
        return "\n\n".join(pages_text)[:15000]
    except Exception:
        return "[Could not extract PDF content]"


def _extract_table(file_path: Path, mime_type: str) -> str:
    """Extract first 50 rows from CSV or Excel and format as markdown table."""
    try:
        if mime_type == "text/csv":
            return _extract_csv(file_path)
        else:
            return _extract_excel(file_path)
    except Exception:
        return "[Could not extract table content]"


def _extract_csv(file_path: Path) -> str:
    """Read CSV and format as markdown table."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = []
        for i, row in enumerate(reader):
            if i >= 51:  # header + 50 rows
                break
            rows.append(row)

    if not rows:
        return "[Empty CSV file]"
    return _rows_to_markdown(rows)


def _extract_excel(file_path: Path) -> str:
    """Read Excel and format as markdown table."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return "[Empty Excel file]"

        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 51:
                break
            rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()

        if not rows:
            return "[Empty Excel file]"
        return _rows_to_markdown(rows)
    except Exception:
        return "[Could not extract Excel content]"


def _extract_document(file_path: Path, mime_type: str) -> str:
    """Extract text from Word documents (.doc/.docx)."""
    try:
        from docx import Document

        doc = Document(str(file_path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)[:15000]
    except Exception:
        return "[Could not extract document content]"


def _rows_to_markdown(rows: list[list[str]]) -> str:
    """Convert rows to a markdown table string."""
    if not rows:
        return ""

    header = rows[0]
    md_lines = []
    md_lines.append("| " + " | ".join(str(h) for h in header) + " |")
    md_lines.append("| " + " | ".join("---" for _ in header) + " |")

    for row in rows[1:]:
        # Pad row if shorter than header
        padded = list(row) + [""] * (len(header) - len(row))
        md_lines.append("| " + " | ".join(str(c) for c in padded[:len(header)]) + " |")

    return "\n".join(md_lines)


def format_attachment_meta(attachment: Attachment) -> dict:
    """Format attachment for API response."""
    return {
        "id": str(attachment.id),
        "filename": attachment.filename,
        "mime_type": attachment.mime_type,
        "size_bytes": attachment.size_bytes,
        "type_category": attachment.type_category,
        "url": f"/api/chat/uploads/{attachment.id}",
    }
